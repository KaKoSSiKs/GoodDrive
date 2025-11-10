from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.permissions import AllowAny
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from rest_framework import status
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
import openpyxl
import logging
from .models import Part, Brand, Warehouse, PartImage

logger = logging.getLogger(__name__)


@api_view(['POST'])
@permission_classes([AllowAny])
@parser_classes([MultiPartParser, FormParser])
def upload_part_image(request):
    """Загрузка изображения товара"""
    try:
        part_id = request.data.get('part_id')
        image_file = request.FILES.get('image')
        
        if not image_file:
            return Response({'error': 'Изображение не предоставлено'}, status=status.HTTP_400_BAD_REQUEST)
        
        if not part_id:
            return Response({'error': 'ID товара не указан'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            part = Part.objects.get(id=part_id)
        except Part.DoesNotExist:
            return Response({'error': 'Товар не найден'}, status=status.HTTP_404_NOT_FOUND)
        
        # Сохраняем изображение
        part_image = PartImage.objects.create(
            part=part,
            image=image_file,
            alt_text=part.title
        )
        
        return Response({
            'id': part_image.id,
            'image_url': part_image.get_image_url,
            'alt_text': part_image.alt_text
        }, status=status.HTTP_201_CREATED)
        
    except Exception as e:
        logger.error(f"Error uploading image: {e}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([AllowAny])
@parser_classes([MultiPartParser, FormParser])
def import_from_excel(request):
    """Импорт товаров из Excel файла"""
    try:
        excel_file = request.FILES.get('file')
        
        if not excel_file:
            return Response({'error': 'Файл не предоставлен'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Проверяем расширение
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            return Response({'error': 'Неверный формат файла. Используйте .xlsx или .xls'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Читаем Excel
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
        
        created_count = 0
        updated_count = 0
        errors = []
        
        # Ожидаемые колонки:
        # A: Название, B: Артикул, C: Бренд, D: Склад, E: Количество, F: Цена, G: Себестоимость
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not row[0]:  # Пропускаем пустые строки
                    continue
                
                title = str(row[0]).strip()
                manufacturer_number = str(row[1]).strip() if row[1] else ''
                brand_name = str(row[2]).strip() if row[2] else 'Без бренда'
                warehouse_name = str(row[3]).strip() if row[3] else 'Основной склад'
                stock = int(row[4]) if row[4] else 0
                price_opt = float(row[5]) if row[5] else 0
                cost_price = float(row[6]) if row[6] and len(row) > 6 else 0
                
                # Получаем или создаём бренд
                brand, _ = Brand.objects.get_or_create(
                    name=brand_name,
                    defaults={'country': 'Не указано'}
                )
                
                # Получаем или создаём склад
                warehouse, _ = Warehouse.objects.get_or_create(
                    name=warehouse_name,
                    defaults={'address': 'Не указано'}
                )
                
                # Проверяем существование товара по артикулу
                if manufacturer_number:
                    part, created = Part.objects.update_or_create(
                        manufacturer_number=manufacturer_number,
                        defaults={
                            'title': title,
                            'brand': brand,
                            'warehouse': warehouse,
                            'stock': stock,
                            'price_opt': price_opt,
                            'cost_price': cost_price,
                            'is_active': True
                        }
                    )
                else:
                    # Если нет артикула, создаём новый товар
                    part = Part.objects.create(
                        title=title,
                        brand=brand,
                        warehouse=warehouse,
                        stock=stock,
                        price_opt=price_opt,
                        cost_price=cost_price,
                        is_active=True
                    )
                    created = True
                
                if created:
                    created_count += 1
                else:
                    updated_count += 1
                    
            except Exception as e:
                errors.append(f"Строка {row_idx}: {str(e)}")
                logger.error(f"Error importing row {row_idx}: {e}")
        
        return Response({
            'status': 'success',
            'created': created_count,
            'updated': updated_count,
            'errors': errors
        })
        
    except Exception as e:
        logger.error(f"Error importing Excel: {e}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([AllowAny])
def download_excel_template(request):
    """Скачать шаблон Excel для импорта"""
    from django.http import HttpResponse
    
    # Создаём новый Excel файл
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Товары"
    
    # Заголовки
    headers = ['Название', 'Артикул', 'Бренд', 'Склад', 'Количество', 'Цена', 'Себестоимость']
    sheet.append(headers)
    
    # Примеры данных
    examples = [
        ['Фильтр масляный', 'OF-123', 'MANN', 'Основной склад', 10, 850.00, 600.00],
        ['Тормозные колодки', 'BP-456', 'BREMBO', 'Склад №2', 5, 2500.00, 1800.00],
        ['Свеча зажигания', 'SP-789', 'NGK', 'Основной склад', 20, 450.00, 300.00],
    ]
    
    for example in examples:
        sheet.append(example)
    
    # Стили для заголовков
    from openpyxl.styles import Font, PatternFill
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Сохраняем в память
    from io import BytesIO
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="template_import_parts.xlsx"'
    
    return response
