"""
Management command для импорта автозапчастей из Excel файла
"""
from django.core.management.base import BaseCommand
from django.db import transaction
from catalog.models import Brand, Warehouse, Part, PartImage
from openpyxl import load_workbook
import os
from decimal import Decimal, InvalidOperation


class Command(BaseCommand):
    help = 'Импортирует автозапчасти из Excel файла'
    
    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='Путь к Excel файлу')
        parser.add_argument(
            '--limit',
            type=int,
            default=100,
            help='Максимальное количество запчастей для импорта (по умолчанию: 100)'
        )
    
    def handle(self, *args, **options):
        file_path = options['file_path']
        limit = options['limit']
        
        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f'Файл {file_path} не найден!'))
            return
        
        self.stdout.write(self.style.SUCCESS(f'Начинаем импорт из файла: {file_path}'))
        self.stdout.write(self.style.SUCCESS(f'Лимит: {limit} запчастей'))
        
        # Загружаем Excel файл
        workbook = load_workbook(file_path, data_only=True)
        worksheet = workbook.active
        
        # Определяем заголовки (предполагаем, что они в первой строке)
        headers = []
        for cell in worksheet[1]:
            headers.append(cell.value.lower() if cell.value else '')
        
        self.stdout.write(f'Найдены колонки: {", ".join(headers)}')
        
        # Счетчики
        brands_created = 0
        warehouses_created = 0
        parts_created = 0
        images_created = 0
        errors = []
        
        # Обрабатываем данные
        with transaction.atomic():
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=False), start=2):
                if row_idx > limit + 1:
                    break
                
                try:
                    # Извлекаем данные из строки
                    row_data = {}
                    for col_idx, cell in enumerate(row):
                        if col_idx < len(headers):
                            row_data[headers[col_idx]] = cell.value
                    
                    # Определяем маппинг полей (настраиваем под вашу структуру Excel)
                    mapping = self.get_field_mapping(row_data)
                    
                    if not mapping:
                        self.stdout.write(self.style.WARNING(f'Пропущена строка {row_idx}: недостаточно данных'))
                        continue
                    
                    # Получаем или создаем бренд
                    brand, created = Brand.objects.get_or_create(
                        name=mapping.get('brand', 'Неизвестный'),
                        defaults={'country': mapping.get('country', '')}
                    )
                    if created:
                        brands_created += 1
                    
                    # Получаем или создаем склад
                    warehouse, created = Warehouse.objects.get_or_create(
                        name=mapping.get('warehouse', 'Основной склад'),
                        defaults={'address': 'Не указан'}
                    )
                    if created:
                        warehouses_created += 1
                    
                    # Создаем автозапчасть
                    part, created = Part.objects.get_or_create(
                        title=mapping['title'],
                        brand=brand,
                        defaults={
                            'label': mapping.get('label', ''),
                            'original_number': mapping.get('original_number', ''),
                            'manufacturer_number': mapping.get('manufacturer_number', ''),
                            'warehouse': warehouse,
                            'quantity': mapping.get('quantity', 0),
                            'stock': mapping.get('stock', 0),
                            'reserve': mapping.get('reserve', 0),
                            'available': mapping.get('available', 0),
                            'price_opt': mapping.get('price_opt', 0),
                            'description': mapping.get('description', ''),
                            'is_active': True
                        }
                    )
                    
                    if created:
                        parts_created += 1
                        
                        # Создаем заглушку для изображения
                        PartImage.objects.get_or_create(
                            part=part,
                            defaults={
                                'image_url': 'https://via.placeholder.com/600x600?text=Auto+Part',
                                'alt_text': mapping['title'],
                                'order_index': 0
                            }
                        )
                        images_created += 1
                    
                except Exception as e:
                    error_msg = f'Ошибка в строке {row_idx}: {str(e)}'
                    errors.append(error_msg)
                    self.stdout.write(self.style.ERROR(error_msg))
                    continue
        
        # Выводим статистику
        self.stdout.write(self.style.SUCCESS('\nИмпорт завершен!'))
        self.stdout.write(self.style.SUCCESS(f'Брендов создано: {brands_created}'))
        self.stdout.write(self.style.SUCCESS(f'Складов создано: {warehouses_created}'))
        self.stdout.write(self.style.SUCCESS(f'Автозапчастей создано: {parts_created}'))
        self.stdout.write(self.style.SUCCESS(f'Изображений создано: {images_created}'))
        
        if errors:
            self.stdout.write(self.style.ERROR(f'\nОшибок: {len(errors)}'))
    
    def get_field_mapping(self, row_data):
        """
        Маппинг данных Excel на модель Part
        Настраиваем под вашу структуру файла
        """
        mapping = {}
        
        # Название товара
        if 'название' in row_data and row_data['название']:
            mapping['title'] = str(row_data['название']).strip()
        elif 'title' in row_data and row_data['title']:
            mapping['title'] = str(row_data['title']).strip()
        else:
            return None  # Название обязательно
        
        # Бренд
        if 'бренд' in row_data:
            mapping['brand'] = str(row_data['бренд']).strip()
        elif 'brand' in row_data:
            mapping['brand'] = str(row_data['brand']).strip()
        else:
            mapping['brand'] = 'Неизвестный'
        
        # Склад
        if 'склад' in row_data:
            mapping['warehouse'] = str(row_data['склад']).strip()
        elif 'warehouse' in row_data:
            mapping['warehouse'] = str(row_data['warehouse']).strip()
        else:
            mapping['warehouse'] = 'Основной склад'
        
        # Страна
        if 'страна' in row_data:
            mapping['country'] = str(row_data['страна']).strip()
        elif 'country' in row_data:
            mapping['country'] = str(row_data['country']).strip()
        else:
            mapping['country'] = ''
        
        # Метка/Марка
        if 'метка' in row_data:
            mapping['label'] = str(row_data['метка']).strip()
        elif 'label' in row_data:
            mapping['label'] = str(row_data['label']).strip()
        else:
            mapping['label'] = ''
        
        # Номера
        if 'оригинальный номер' in row_data:
            mapping['original_number'] = str(row_data['оригинальный номер']).strip()
        elif 'original_number' in row_data:
            mapping['original_number'] = str(row_data['original_number']).strip()
        else:
            mapping['original_number'] = ''
        
        if 'номер производителя' in row_data:
            mapping['manufacturer_number'] = str(row_data['номер производителя']).strip()
        elif 'manufacturer_number' in row_data:
            mapping['manufacturer_number'] = str(row_data['manufacturer_number']).strip()
        else:
            mapping['manufacturer_number'] = ''
        
        # Количества
        mapping['quantity'] = self.parse_int(row_data.get('количество') or row_data.get('quantity'), 0)
        mapping['stock'] = self.parse_int(row_data.get('на складе') or row_data.get('stock'), 0)
        mapping['reserve'] = self.parse_int(row_data.get('в резерве') or row_data.get('reserve'), 0)
        mapping['available'] = mapping['stock'] - mapping['reserve']  # Вычисляем доступное
        
        # Цена
        mapping['price_opt'] = self.parse_decimal(
            row_data.get('цена') or row_data.get('price') or row_data.get('price_opt'), 
            0
        )
        
        # Описание
        if 'описание' in row_data:
            mapping['description'] = str(row_data['описание']).strip()
        elif 'description' in row_data:
            mapping['description'] = str(row_data['description']).strip()
        else:
            mapping['description'] = ''
        
        return mapping
    
    def parse_int(self, value, default=0):
        """Преобразует значение в integer"""
        if value is None:
            return default
        try:
            if isinstance(value, (int, float)):
                return int(value)
            return int(str(value).strip())
        except (ValueError, AttributeError):
            return default
    
    def parse_decimal(self, value, default=0):
        """Преобразует значение в Decimal"""
        if value is None:
            return default
        try:
            if isinstance(value, Decimal):
                return value
            return Decimal(str(value).strip().replace(',', '.'))
        except (ValueError, InvalidOperation):
            return default
